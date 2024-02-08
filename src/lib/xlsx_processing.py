# Desc: xlsx processing functions
from datetime import datetime
from openpyxl import Workbook, load_workbook
from lib.config_manager import load_config
import os
import shutil

config = load_config()
xlsx_filepath = config['xlsx_file']
xlsx_sheet_name = config['xlsx_sheet_name']
xlsx_columns = config['xlsx_columns']


# open xlsx file
wb = Workbook()
wb = load_workbook(filename=xlsx_filepath)


def get_cell_by_email(email):
    sheet = wb[xlsx_sheet_name]
    names_row = sheet[1]
    email_column_names = xlsx_columns['email_columns']
    for email_column_name in email_column_names:
        email_column_index = None
        for cell in names_row:
            if cell.value == email_column_name:
                email_column_index = cell.column
                break
        if email_column_index is not None:
            for row in sheet.iter_rows(min_row=2, min_col=email_column_index, max_col=email_column_index):
                for cell in row:
                    if cell.value and email:
                        if cell.value.strip().lower() == email.strip().lower():
                            return cell
    return None


def get_id(email):
    # get value of column xlsx_columns['email_columns']['id'] of row with email
    sheet = wb[xlsx_sheet_name]
    names_row = sheet[1]
    id_column_index = 0
    # email_column_names_are_strings = xlsx_columns['email_columns']
    for cell in names_row:
        if cell.value == xlsx_columns['id']:
            id_column_index = cell.column

    email_cell = get_cell_by_email(email)
    if email_cell is not None:
        id_cell = sheet.cell(row=email_cell.row, column=id_column_index)
        return id_cell.value
    else:
        return None


def save_xlsx_file():
    xlsx_file_path = os.path.dirname(xlsx_filepath)
    backup_dir = f'{xlsx_file_path}/backup'
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = f'{backup_dir}/{os.path.basename(xlsx_filepath)}_{timestamp}'
    if shutil.copyfile(xlsx_filepath, backup_file):
        wb.save(xlsx_filepath)
    else:
        print('Error: could not save backup file. Skipping saving xlsx file.')


def set_if_replied(email):
    # set column replied to true, if email is found either in column email or column reply_from
    sheet = wb[xlsx_sheet_name]
    names_row = sheet[1]
    replied_column_index = 0
    # email_column_names_are_strings = xlsx_columns['email_columns']
    for cell in names_row:
        if cell.value == xlsx_columns['replied']:
            replied_column_index = cell.column

    email_cell = get_cell_by_email(email)
    if email_cell is not None:
        replied_cell = sheet.cell(
            row=email_cell.row, column=replied_column_index)
        replied_cell.value = True
        save_xlsx_file()
        return True
    else:
        return False


def update_excel_file(emails):
    # update excel file with new data

    # check that the excel file not not currently opened in the windows programm Excel. check by testing for the existence of the file starting with a tilde
    filename = os.path.basename(xlsx_filepath)
    directory = os.path.dirname(xlsx_filepath)

    if os.path.exists(os.path.join(directory, f'~${filename}')):
        print(
            f'Error: Updating the Excel file failed. Excel file {xlsx_filepath} is currently open in Excel. Please close the file and try again.')
        return

    for email in emails:
        set_if_replied(email)
        # if not set_if_replied(email):
        #     print(f'Warning: email {email} not found in xlsx file')
